from io import BytesIO
from pathlib import Path

from fastmcp.utilities.logging import get_logger
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import (
    coordinate_from_string,
    coordinate_to_tuple,
    get_column_letter,
    range_boundaries,
)
from openpyxl.utils.exceptions import CellCoordinatesException
from openpyxl.workbook import Workbook as WorkbookType
from openpyxl.worksheet.worksheet import Worksheet

from models.xlsx import SheetInfo

logger = get_logger(__name__)


_MAX_COLUMN_WIDTH = 60.0

_CHART_TYPES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
}

_COMMENT_AUTHOR = "Assistant"

# Inserted pictures are capped to this width in pixels.
_MAX_PICTURE_WIDTH = 640


def list_template_names(
    templates_dir: Path,
) -> list[str]:

    if not templates_dir.is_dir():
        raise RuntimeError(f"Templates directory not found: {templates_dir}")

    template_names: list[str] = []

    for file_path in sorted(templates_dir.glob("*.xlsx")):
        try:
            load_workbook(file_path, read_only=True).close()
        except Exception as error:
            logger.warning(f"Skipping template {file_path.name}: {error}")
        else:
            template_names.append(file_path.name)

    return template_names


def count_sheets(
    workbook: WorkbookType,
) -> int:
    return len(workbook.worksheets)


def list_sheet_infos(
    workbook: WorkbookType,
) -> list[SheetInfo]:

    sheet_infos: list[SheetInfo] = []

    for worksheet in workbook.worksheets:

        rows, cols = worksheet.max_row, worksheet.max_column

        # openpyxl never reports below 1x1, so treat a lone empty A1 as empty.
        if rows == 1 and cols == 1 and worksheet["A1"].value is None:
            rows, cols = 0, 0

        sheet_infos.append(
            SheetInfo(title=worksheet.title, rows=rows, cols=cols)
        )

    return sheet_infos


def _get_sheet(
    workbook: WorkbookType,
    sheet_title: str,
) -> Worksheet:
    try:
        return workbook[sheet_title]
    except KeyError:
        raise ValueError(
            f"Sheet '{sheet_title}' not found."
        ) from None


def _require_ref(
    ref: str,
) -> None:
    try:
        coordinate_from_string(ref)
    except CellCoordinatesException:
        raise ValueError(f"Invalid cell reference '{ref}'.") from None


def _range_reference(
    worksheet: Worksheet,
    ref: str,
) -> Reference:

    try:
        boundaries = range_boundaries(ref)
    except ValueError:
        raise ValueError(f"Invalid range reference '{ref}'.") from None

    if None in boundaries:
        raise ValueError(
            f"Range '{ref}' is unbounded. Use a bounded range like 'B1:C4'."
        )

    min_col, min_row, max_col, max_row = boundaries

    return Reference(
        worksheet,
        min_col=min_col, min_row=min_row,
        max_col=max_col, max_row=max_row,
    )


def read_sheet(
    workbook: WorkbookType,
    sheet_title: str,
) -> list[list[str]]:

    worksheet = _get_sheet(workbook, sheet_title)

    rows: list[list[str]] = []

    for row in worksheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value) for value in row])

    return rows


def write_rows(
    workbook: WorkbookType,
    sheet_title: str,
    rows: list[list[bool | int | float | str | None]],
    anchor: str,
    style: str | None,
) -> None:

    worksheet = _get_sheet(workbook, sheet_title)

    try:
        start_row, start_col = coordinate_to_tuple(anchor)
    except (CellCoordinatesException, ValueError):
        raise ValueError(f"Invalid anchor reference '{anchor}'.") from None

    if style is not None and style not in workbook.named_styles:
        raise ValueError(f"Style '{style}' not found.")

    for row_offset, row in enumerate(rows):
        for col_offset, value in enumerate(row):

            cell = worksheet.cell(
                row=start_row + row_offset,
                column=start_col + col_offset,
            )
            cell.value = value

            if style is not None:
                cell.style = style


def write_cell(
    workbook: WorkbookType,
    sheet_title: str,
    ref: str,
    value: bool | int | float | str | None,
    style: str | None,
) -> None:

    worksheet = _get_sheet(workbook, sheet_title)

    _require_ref(ref)

    if style is not None and style not in workbook.named_styles:
        raise ValueError(f"Style '{style}' not found.")

    cell = worksheet[ref]
    cell.value = value

    if style is not None:
        cell.style = style


def insert_picture(
    workbook: WorkbookType,
    sheet_title: str,
    image: BytesIO,
    anchor: str,
) -> None:

    worksheet = _get_sheet(workbook, sheet_title)

    _require_ref(anchor)

    picture = Image(image)

    # openpyxl re-reads and closes the image buffer on every save; pin the
    # bytes once so the workbook survives a second `finalize_project`.
    data = picture._data()
    picture._data = lambda: data

    if picture.width > _MAX_PICTURE_WIDTH:
        picture.height = round(
            picture.height * _MAX_PICTURE_WIDTH / picture.width
        )
        picture.width = _MAX_PICTURE_WIDTH

    worksheet.add_image(picture, anchor)


def add_chart(
    workbook: WorkbookType,
    sheet_title: str,
    kind: str,
    data: str,
    categories: str,
    title: str | None,
    anchor: str,
) -> None:

    worksheet = _get_sheet(workbook, sheet_title)

    if kind not in _CHART_TYPES:
        raise ValueError(
            f"Chart kind '{kind}' not supported. "
            f"Use one of: {', '.join(_CHART_TYPES)}."
        )

    _require_ref(anchor)

    data_reference = _range_reference(worksheet, data)
    categories_reference = _range_reference(worksheet, categories)

    if kind == "pie" and data_reference.min_col != data_reference.max_col:
        raise ValueError("A pie chart takes exactly one series column.")

    # `data` starts with the series-name row; `categories` must label
    # exactly the value rows below it.
    if (
        categories_reference.min_row != data_reference.min_row + 1
        or categories_reference.max_row != data_reference.max_row
    ):
        raise ValueError(
            f"categories '{categories}' must label the data's value rows "
            f"{data_reference.min_row + 1}-{data_reference.max_row} "
            f"(data '{data}' without its series-name row)."
        )

    chart = _CHART_TYPES[kind]()
    chart.add_data(data_reference, titles_from_data=True)
    chart.set_categories(categories_reference)

    if title is not None:
        chart.title = title

    worksheet.add_chart(chart, anchor)


def add_comment(
    workbook: WorkbookType,
    sheet_title: str,
    ref: str,
    text: str,
) -> None:

    worksheet = _get_sheet(workbook, sheet_title)

    _require_ref(ref)

    worksheet[ref].comment = Comment(text, _COMMENT_AUTHOR)


def clear_workbook(
    workbook: WorkbookType,
) -> None:

    keep = workbook.create_sheet()

    for worksheet in list(workbook.worksheets):
        if worksheet is not keep:
            workbook.remove(worksheet)

    keep.title = "Sheet"


def add_sheet(
    workbook: WorkbookType,
    title: str,
    index: int | None,
) -> None:

    if title in workbook.sheetnames:
        raise ValueError(f"Sheet '{title}' already exists.")

    workbook.create_sheet(title=title, index=index)


def remove_sheets(
    workbook: WorkbookType,
    titles: list[str],
) -> None:

    targets = [_get_sheet(workbook, title) for title in dict.fromkeys(titles)]

    if len(targets) >= len(workbook.worksheets):
        raise ValueError(
            "Cannot remove every sheet; a workbook needs at least one."
        )

    for worksheet in targets:
        workbook.remove(worksheet)


def move_sheet(
    workbook: WorkbookType,
    title: str,
    to_index: int,
) -> None:

    try:
        current = workbook.sheetnames.index(title)
    except ValueError:
        raise ValueError(
            f"Sheet '{title}' not found."
        ) from None

    count = len(workbook.sheetnames)

    if not -count <= to_index < count:
        raise ValueError(f"Target index {to_index} out of range.")

    # openpyxl moves by relative offset; `% count` maps negative targets.
    workbook.move_sheet(title, offset=to_index % count - current)


def autofit_columns(
    workbook: WorkbookType,
) -> None:

    for worksheet in workbook.worksheets:
        for index, column in enumerate(worksheet.iter_cols(), start=1):

            lengths = [
                len(line)
                for cell in column
                if cell.value is not None
                for line in str(cell.value).splitlines()
            ]

            if lengths:
                letter = get_column_letter(index)
                worksheet.column_dimensions[letter].width = min(
                    max(lengths) + 2.0, _MAX_COLUMN_WIDTH,
                )

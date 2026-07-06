from asyncio import to_thread
from io import BytesIO
from zipfile import BadZipFile

from fastmcp import FastMCP
from fastmcp.dependencies import CurrentAccessToken, Depends, TokenClaim
from fastmcp.exceptions import ToolError
from fastmcp.server.auth import AccessToken
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import Field

from config import get_settings
from models._base import TemplatesResult
from models.xlsx import (
    CellInput,
    FinalizeResult,
    Project,
    ProjectResult,
    ReadResult,
    SheetsResult,
    StylesResult,
    WriteResult,
)
from services.owui import download_file, upload_file
from subservers._store import ProjectStore
from subservers.xlsx._utils import (
    clear_workbook,
    count_sheets,
    drop_sheets,
    insert_sheet as _insert_sheet,
    list_sheet_infos,
    list_template_names,
    move_sheet as _move_sheet,
    read_sheet_rows,
    write_cells as _write_cells,
    write_rows as _write_rows,
)

XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_EDIT_HINT = (
    "Continue the edit batch; when the user's request is fully applied, "
    "call `finalize_project` once."
)

_FINALIZE_HINT = (
    "Share `owui_url` with the user as the download link. The project stays "
    "active; end a later edit batch with `finalize_project` again."
)


_settings = get_settings()

_store = ProjectStore[Project](
    max_size=1_000,
    ttl=_settings.project_ttl,
    sweep_interval=_settings.project_sweep_interval,
)

def _get_project(
    user_id: str = TokenClaim("id"),
) -> Project:

    project = _store.get(user_id)

    if project is None:
        # ToolError passes dependency resolution unchanged; other exceptions
        # get swallowed into a generic "failed to resolve" message.
        raise ToolError(
            "No project for this user. "
            "Call `create_project` or `open_project` first."
        )

    return project


mcp = FastMCP(name="xlsx", lifespan=_store.lifespan)


@mcp.tool(
    name="list_templates",
    description=(
        "List the available templates. Use this when the user did NOT attach "
        "a file, then pick one for `create_project`. If the user attached a "
        "file, use `open_project` instead."
    ),
)
async def list_templates() -> TemplatesResult:

    templates = await to_thread(list_template_names, _settings.templates_dir)

    hint = (
        "Pick a template and call `create_project`."
    ) if templates else (
        "No templates available. Ask the administrator to add `.xlsx` templates."
    )

    return TemplatesResult(hint=hint, templates=templates)


@mcp.tool(
    name="create_project",
    description=(
        "Create a new, empty in-memory project from a template. Use this by "
        "default when the user did NOT attach a file. Overwrites any existing "
        "project for the user. Then call `list_sheets`."
    ),
)
async def create_project(
    template_name: str = Field(description="From `list_templates`."),
    user_id: str = TokenClaim("id"),
) -> ProjectResult:

    templates = await to_thread(list_template_names, _settings.templates_dir)

    if template_name not in templates:
        raise ValueError(
            f"Template '{template_name}' not found. "
            "Pick one from `list_templates`."
        )

    template_file = _settings.templates_dir.joinpath(template_name)

    workbook = await to_thread(load_workbook, template_file)
    clear_workbook(workbook)

    _store.set(user_id, Project(workbook=workbook))

    return ProjectResult(
        hint=(
            f"Empty project created from template '{template_name}' with one "
            "sheet 'Sheet'. Call `list_styles` for the available styles, then "
            "`write_rows` / `write_cells` to fill it."
        ),
        sheet_count=count_sheets(workbook),
    )


@mcp.tool(
    name="open_project",
    description=(
        "Open a `.xlsx` the user attached in OpenWebUI, by its `file_id`. Use "
        "only when the user actually attached a file; if none was given, use "
        "`create_project` instead. Overwrites any existing project for the "
        "user. Then call `list_sheets` to inspect or extend it."
    ),
)
async def open_project(
    file_id: str = Field(
        description=(
            "OpenWebUI file ID of the file the user attached. NOT a template "
            "name from `list_templates`, and never invented."
        ),
    ),
    token: AccessToken = CurrentAccessToken(),
    user_id: str = TokenClaim("id"),
) -> ProjectResult:

    file_name, file_content = await download_file(
        file_id=file_id,
        token=token.token,
    )

    try:
        workbook = await to_thread(load_workbook, BytesIO(file_content))
    except (BadZipFile, InvalidFileException) as error:
        raise ValueError(
            f"File '{file_id}' is not a valid `.xlsx` workbook."
        ) from error

    _store.set(user_id, Project(workbook=workbook))

    return ProjectResult(
        hint=(
            f"Project opened from attached file '{file_name}'. "
            "Call `list_sheets` to see its sheets, then `read_sheet` or the "
            "write tools to work with them."
        ),
        sheet_count=count_sheets(workbook),
    )


@mcp.tool(
    name="list_sheets",
    description=(
        "List the worksheets of the current project with their used extent "
        "(`rows` x `cols`). Use the titles to target the other tools."
    ),
)
async def list_sheets(
    project: Project = Depends(_get_project),
) -> SheetsResult:

    async with project.lock:
        sheets = list_sheet_infos(project.workbook)

    return SheetsResult(
        hint=(
            "Use the sheet titles with `read_sheet`, `write_rows`, and "
            "`write_cells`."
        ),
        sheets=sheets,
    )


@mcp.tool(
    name="list_styles",
    description=(
        "List the names of the named cell styles in the current project. Pass "
        "one to `write_cells` or `write_rows` to format cells. Optional — "
        "cells without a style use the default formatting."
    ),
)
async def list_styles(
    project: Project = Depends(_get_project),
) -> StylesResult:

    async with project.lock:
        styles = list(project.workbook.named_styles)

    hint = (
        "Pass a style name to `write_rows` or `write_cells`; omit it for "
        "default formatting."
    ) if styles else (
        "No named styles in this workbook; write without `style`."
    )

    return StylesResult(hint=hint, styles=styles)


@mcp.tool(
    name="insert_sheet",
    description=(
        "Insert an empty worksheet. Without `index` it is appended; otherwise "
        "it is inserted at that zero-based position. After the requested "
        "batch of edits, call `finalize_project` once (not after each change)."
    ),
)
async def insert_sheet(
    title: str = Field(description="Title for the new worksheet."),
    index: int | None = Field(
        default=None,
        description=(
            "Zero-based position to insert at. If omitted, the sheet is "
            "appended at the end."
        ),
    ),
    user_id: str = TokenClaim("id"),
    project: Project = Depends(_get_project),
) -> ProjectResult:

    async with project.lock:

        _insert_sheet(project.workbook, title, index)

        _store.touch(user_id, project)

        return ProjectResult(
            hint=_EDIT_HINT,
            sheet_count=count_sheets(project.workbook),
        )


@mcp.tool(
    name="write_rows",
    description=(
        "Fill a contiguous block of cells from `rows` (row-major), starting "
        "at `anchor` (top-left, default `A1`). Best for tables and bulk data; "
        "for scattered or individually-styled cells use `write_cells`. An "
        "optional `style` from `list_styles` formats every written cell. "
        "Changes stay in memory only. After the requested batch of edits, "
        "call `finalize_project` once (not after each change)."
    ),
)
async def write_rows(
    sheet: str = Field(description="Worksheet title from `list_sheets`."),
    rows: list[list[bool | int | float | str | None]] = Field(
        description=(
            "Row-major values. Numbers stay numeric, a string starting with "
            "`=` becomes a formula, and `null` clears the cell."
        ),
    ),
    anchor: str = Field(
        default="A1",
        description="A1 reference of the top-left cell.",
    ),
    style: str | None = Field(
        default=None,
        description="Named cell style from `list_styles` for every cell.",
    ),
    user_id: str = TokenClaim("id"),
    project: Project = Depends(_get_project),
) -> WriteResult:

    async with project.lock:

        _write_rows(project.workbook, sheet, anchor, rows, style)

        _store.touch(user_id, project)

    return WriteResult(
        hint=_EDIT_HINT,
        cells=sum(len(row) for row in rows),
    )


@mcp.tool(
    name="write_cells",
    description=(
        "Write values into individual cells by A1 reference, optionally "
        "formatting each with a named style from `list_styles`. Best for "
        "scattered edits or styling specific cells; to fill a contiguous "
        "table use `write_rows`. At most a limited number of cells per call. "
        "Only the listed cells change; the rest stay as-is. Changes stay in "
        "memory only. "
        "After the requested batch of edits, call `finalize_project` once "
        "(not after each change)."
    ),
)
async def write_cells(
    sheet: str = Field(description="Worksheet title from `list_sheets`."),
    cells: list[CellInput] = Field(
        max_length=100,
        description="Cells to write, each targeting one A1 reference.",
    ),
    user_id: str = TokenClaim("id"),
    project: Project = Depends(_get_project),
) -> WriteResult:

    async with project.lock:

        _write_cells(project.workbook, sheet, cells)

        _store.touch(user_id, project)

    return WriteResult(hint=_EDIT_HINT, cells=len(cells))


@mcp.tool(
    name="read_sheet",
    description=(
        "Read a worksheet's used range as rows of plain text (row-major, empty "
        "cells as ``). Formulas are returned as their formula string, not the "
        "computed result."
    ),
)
async def read_sheet(
    sheet: str = Field(description="Worksheet title from `list_sheets`."),
    project: Project = Depends(_get_project),
) -> ReadResult:

    async with project.lock:
        rows = read_sheet_rows(project.workbook, sheet)

    hint = (
        "Row and column positions are zero-based offsets from `A1`; use A1 "
        "references with the write tools."
    ) if rows else (
        "Sheet is empty. Fill it with `write_rows`."
    )

    return ReadResult(hint=hint, rows=rows)


@mcp.tool(
    name="move_sheet",
    description=(
        "Move a worksheet to a new position by zero-based index. Negative "
        "`to_index` counts from the end. Changes stay in memory only. After "
        "the requested batch of edits, call `finalize_project` once (not "
        "after each change)."
    ),
)
async def move_sheet(
    title: str = Field(description="Worksheet title from `list_sheets`."),
    to_index: int = Field(description="Zero-based target position."),
    user_id: str = TokenClaim("id"),
    project: Project = Depends(_get_project),
) -> SheetsResult:

    async with project.lock:

        _move_sheet(project.workbook, title, to_index)

        _store.touch(user_id, project)

        return SheetsResult(
            hint=_EDIT_HINT,
            sheets=list_sheet_infos(project.workbook),
        )


@mcp.tool(
    name="remove_sheets",
    description=(
        "Remove worksheets by title (from `list_sheets`). A workbook must keep "
        "at least one sheet. Changes stay in memory only. After the requested "
        "batch of edits, call `finalize_project` once (not after each change)."
    ),
)
async def remove_sheets(
    titles: list[str] = Field(description="Worksheet titles to remove."),
    user_id: str = TokenClaim("id"),
    project: Project = Depends(_get_project),
) -> ProjectResult:

    async with project.lock:

        drop_sheets(project.workbook, titles)

        _store.touch(user_id, project)

        return ProjectResult(
            hint=_EDIT_HINT,
            sheet_count=count_sheets(project.workbook),
        )


@mcp.tool(
    name="finalize_project",
    description=(
        "Final step of an edit batch: serialize the current project to "
        "`.xlsx` and upload it to OpenWebUI. Call exactly once after the "
        "requested batch of changes, not after each one. The project stays "
        "active afterwards, so a later edit batch ends with another single "
        "`finalize_project` call."
    ),
)
async def finalize_project(
    file_name: str = Field(
        min_length=1, max_length=60,
        description="Stem without `.xlsx`.",
    ),
    token: AccessToken = CurrentAccessToken(),
    user_id: str = TokenClaim("id"),
    project: Project = Depends(_get_project),
) -> FinalizeResult:

    async with project.lock:

        out_name = f"{file_name}.xlsx"

        buffer = BytesIO()
        await to_thread(project.workbook.save, buffer)

        _store.touch(user_id, project)

        sheet_count = count_sheets(project.workbook)

    uploaded = await upload_file(
        file_name=out_name,
        data=buffer.getvalue(),
        content_type=XLSX_MIME,
        token=token.token,
    )

    return FinalizeResult(
        hint=_FINALIZE_HINT,
        file_name=out_name,
        sheet_count=sheet_count,
        owui_url=uploaded.download_url,
    )
